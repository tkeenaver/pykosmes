import openpyxl
from openpyxl.styles import Alignment
import os

#워크북(Workbook) 객체 만들기
wb = openpyxl.Workbook()   

#시트(Sheet) 객체 만들기
ws = wb.create_sheet(index=0, title='Merge')
wb.remove(wb['Sheet'])

#데이터 입력하기
tuple_of_rows = ((1, 2),
                 (3, 4), 
                 (5, 6),
                 (7, 8),
                 (9, 10),
                 )

for row in tuple_of_rows:
    ws.append(row)
    print(row)

ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
A1 = ws.cell(row=1, column=1)
A1.value = "Merged"
A1.alignment = Alignment(horizontal='center', vertical='center')

#워크북의 변경내용을 새로운 파일에 저장
wb.save(os.path.join(os.getcwd(), "output", "create_workbook3.xlsx"))
print("\n*** output 폴더의 \"create_workbook3.xlsx\" 파일을 확인해 보세요. ***\n")
