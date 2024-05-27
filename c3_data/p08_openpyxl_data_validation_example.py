import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active

ws['A1'] = '등록번호'
ws['B1'] = '동물종류'
for i in range(2,12):
    ws.cell(row=i, column=1).value = str(i+100)

dv = DataValidation(type='list', formula1='"강아지,고양이,햄스터"', showDropDown=False)
ws.add_data_validation(dv)
dv.add('B2:B11')

for c_cells in ws.columns:
    new_c_len = max(len(str(cell.value)) for cell in c_cells)
    new_c_let = (get_column_letter(c_cells[0].column))
    if new_c_len > 0:
        ws.column_dimensions[new_c_let].width = new_c_len*2.5  # 2.5 for dbcs

file_path = "output/data_validation.xlsx"
wb.save(file_path)
