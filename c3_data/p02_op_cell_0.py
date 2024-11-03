import openpyxl
import openpyxl.utils.cell as ut_cell
import os

cwd = os.getcwd()
filename = "df.xlsx"
filepath = os.path.join(cwd, "data", filename)

wb = openpyxl.load_workbook(filepath)   
ws = wb['Sheet1']

b2 = ws['B2']
print(b2, b2.coordinate, b2.column, b2.row, b2.value)
print(ut_cell.get_column_letter(b2.column))
print(ut_cell.column_index_from_string('AB'))

c4 = ws.cell(row=4, column=3)
print(c4, c4.coordinate, c4.value)

b2.value = 50
c4.value = 60
print(f'{b2.value=}')
print(f'{c4.value=}')

wb.save(os.path.join(cwd, "output", "df2.xlsx"))
