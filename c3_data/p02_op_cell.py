import openpyxl
import openpyxl.utils.cell as ut_cell
import os

cwd = os.getcwd()
filename = "df.xlsx"
filepath = os.path.join(cwd, "data", filename)

wb = openpyxl.load_workbook(filepath)   
ws = wb['Sheet1']

#셀(Cell) 객체
b2 = ws['B2']
print(b2, b2.coordinate, b2.column, b2.row, b2.value)
print(ut_cell.get_column_letter(b2.column))
print(ut_cell.column_index_from_string('AB'))

c4 = ws.cell(row=4, column=3)
print(c4, c4.coordinate, c4.value)

b2.value = 50
c4.value = 60
print(f'{b2.value=}')
print(f'{c4.value=}')

for row in ws.iter_rows(min_row=1, max_col=3, max_row=4, values_only=True):
    for cell in row:
        print(cell,end = ',')
    print()
ws.append(['A', 3.14, 'C'])

from openpyxl.chart import BarChart, Reference
chart = BarChart()
data = Reference(ws, min_col=1, min_row=1, max_col=3, max_row=4)
chart.add_data(data, titles_from_data=True)
ws.add_chart(chart, 'E5')

from openpyxl.drawing.image import Image
img = Image('data/boxplot.png')
ws.add_image(img, 'A20')

#워크북의 변경내용을 새로운 파일에 저장
wb.save(os.path.join(cwd, "output", "df2.xlsx"))
print("\n*** output 폴더의 \"df2.xlsx\" 파일을 확인해 보세요. ***\n")
