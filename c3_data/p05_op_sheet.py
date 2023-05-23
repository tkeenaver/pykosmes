import openpyxl
import openpyxl.utils.cell as ut_cell
from openpyxl.styles import Alignment, Font
import os

#워크북(Workbook) 객체 만들기
wb = openpyxl.Workbook()   

#시트(Sheet) 객체 만들기
ws1 = wb.create_sheet(index=0, title='Column')
print(wb.sheetnames)
wb.remove(wb['Sheet'])
print(wb.sheetnames)
ws2 = wb.create_sheet(index=1, title='Row')
print(wb.sheetnames)

for col in ws1.iter_cols(min_row=1, min_col=1, max_row=3, max_col=6):
    print(col)
    for each_cell in col:   
        each_cell.value = ut_cell.get_column_letter(each_cell.column)
        each_cell.alignment = Alignment(horizontal='right', vertical='center')
        each_cell.font = Font(bold=True, name='Arial', size=12, underline='single', color='1bb638')

for row in ws2.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    print(row)
    for each_cell in row:   
        each_cell.value = each_cell.row
        each_cell.alignment = Alignment(horizontal='center', vertical='center')
        each_cell.font = Font(italic=True, name='Consoras', size=10, color='ff0000')

#워크북의 변경내용을 새로운 파일에 저장
wb.save(os.path.join(os.getcwd(), "output", "create_workbook2.xlsx"))
print("\n*** output 폴더의 \"create_workbook2.xlsx\" 파일을 확인해 보세요. ***\n")
