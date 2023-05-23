import openpyxl
from openpyxl.drawing.image import Image
import os

cwd = os.getcwd()
filename = "boxplot.png"
filepath = os.path.join(cwd, "data", filename)
img = Image(filepath)
#print(img)

wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0, title='Image')
wb.remove(wb['Sheet'])

ws['A1'] = "Box Plot"
ws.add_image(img, 'B3')

wb.save(os.path.join(os.getcwd(), "output", "insert_image.xlsx"))
print("\n*** output 폴더의 \"insert_image.xlsx\" 파일을 확인해 보세요. ***\n")
