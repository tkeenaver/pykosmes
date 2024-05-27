import openpyxl
import openpyxl.utils.cell as ut_cell
import os

cwd = os.getcwd()
filename = "df.xlsx"
filepath = os.path.join(cwd, "data", filename)

wb = openpyxl.load_workbook(filepath)   
ws = wb['Sheet1']

from openpyxl.styles import Font, Alignment
ws['A1'].font = Font(name='Arial', size=12, bold=True, color='FF0000')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

from openpyxl.styles import PatternFill
ws['A1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

wb.save(os.path.join(cwd, "output", "df3.xlsx"))
print("\n*** output 폴더의 \"df3.xlsx\" 파일을 확인해 보세요. ***\n")
